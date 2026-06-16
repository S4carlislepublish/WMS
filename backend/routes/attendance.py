from flask import Blueprint, request, jsonify
from models.database import db
from models.attendance import Attendance
from datetime import datetime
from models.employee import Employee
from models.user import User
from datetime import date
from sqlalchemy import extract
from datetime import timedelta
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from flask import send_file

from io import BytesIO


from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter


attendance_bp = Blueprint(
    "attendance",
    __name__
)


@attendance_bp.route("/checkin", methods=["POST"])
def check_in():

    try:

        data = request.json

        user_id = data.get("user_id")

        if not user_id:
            return jsonify({
                "success": False,
                "message": "User ID is required"
            }), 400

        employee = Employee.query.filter_by(
            user_id=user_id
        ).first()

        if not employee:
            return jsonify({
                "success": False,
                "message": "Employee not found"
            }), 404

        # =====================================
        # SHIFT VALIDATION
        # =====================================

        shift_name = (
            employee.shift_timing or ""
        ).strip().lower()

        current_time = datetime.now().time()
        
        print("================================")
        print("Employee Shift:", employee.shift_timing)
        print("Current Time:", current_time)
        print("================================")

        # First Shift (06:00 AM - 02:00 PM)
        if shift_name == "first shift":
            print("================================")
            print("USER ID:", user_id)
            print("EMPLOYEE:", employee.first_name)
            print("SHIFT RAW:", employee.shift_timing)
            print("SHIFT LOWER:", shift_name)
            print("CURRENT TIME:", current_time)
            print("================================")

            allowed_time = datetime.strptime(
                "06:00",
                "%H:%M"
            ).time()

            if current_time < allowed_time:

                return jsonify({
                    "success": False,
                    "message":
                    "First Shift check-in allowed only after 06:00 AM"
                }), 400

        # General Shift (09:00 AM - 06:00 PM)
        elif shift_name == "general shift":

            allowed_time = datetime.strptime(
                "09:00",
                "%H:%M"
            ).time()

            if current_time < allowed_time:

                return jsonify({
                    "success": False,
                    "message":
                    "General Shift check-in allowed only after 09:00 AM"
                }), 400

        # Second Shift (02:00 PM - 10:00 PM)
        elif shift_name == "second shift":

            allowed_time = datetime.strptime(
                "14:00",
                "%H:%M"
            ).time()

            if current_time < allowed_time:

                return jsonify({
                    "success": False,
                    "message":
                    "Second Shift check-in allowed only after 02:00 PM"
                }), 400

        # Night Shift (10:00 PM - 06:00 AM)
        elif shift_name == "night shift":

            allowed_time = datetime.strptime(
                "22:00",
                "%H:%M"
            ).time()


            if current_time < allowed_time:

                return jsonify({
                    "success": False,
                    "message":
                    "Night Shift check-in allowed only after 10:00 PM"
                }), 400

        # =====================================
        # CHECK ALREADY CHECKED IN
        # =====================================

        today = datetime.now().date()

        attendance = Attendance.query.filter_by(
            user_id=user_id,
            attendance_date=today
        ).first()

        if attendance:

            return jsonify({
                "success": False,
                "message":
                "You have already checked in today."
            }), 400

        # =====================================
        # CREATE ATTENDANCE
        # =====================================

        attendance = Attendance(
            user_id=user_id,
            attendance_date=today,
            check_in=datetime.now(),
            status="Present"
        )

        db.session.add(attendance)

        db.session.commit()

        return jsonify({
            "success": True,
            "message":
            "Checked In Successfully"
        })

    except Exception as e:

        db.session.rollback()

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@attendance_bp.route("/checkout", methods=["POST"])
def check_out():

    try:

        data = request.json

        user_id = data.get("user_id")

        if not user_id:
            return jsonify({
                "success": False,
                "error": "User ID is required"
            }), 400

        print("CHECKOUT USER ID:", user_id)

        attendance = Attendance.query.filter(
            Attendance.user_id == user_id,
            Attendance.check_in.isnot(None),
            Attendance.check_out.is_(None)
        ).order_by(
            Attendance.id.desc()
        ).first()

        print("ATTENDANCE FOUND:", attendance)

        if not attendance:
            return jsonify({
                "success": False,
                "error": "No active check-in found"
            }), 404

        if not attendance.check_in:
            return jsonify({
                "success": False,
                "error": "Check-in time missing"
            }), 400

        attendance.check_out = datetime.now()

        total_seconds = (
            attendance.check_out -
            attendance.check_in
        ).total_seconds()

        break_minutes = (
            attendance.total_break_minutes or 0
        )

        total_seconds -= break_minutes * 60

        attendance.total_hours = round(
            total_seconds / 3600,
            2
        )

        attendance.status = "Present"

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Checked Out Successfully",
            "check_in": attendance.check_in.strftime("%Y-%m-%d %H:%M:%S"),
            "check_out": attendance.check_out.strftime("%Y-%m-%d %H:%M:%S"),
            "total_hours": attendance.total_hours
        }), 200

    except Exception as e:

        db.session.rollback()

        print("CHECKOUT ERROR:", str(e))

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@attendance_bp.route("/status/<int:user_id>")
def attendance_status(user_id):

    attendance = Attendance.query.filter_by(
        user_id=user_id,
        check_out=None
    ).order_by(
        Attendance.id.desc()
    ).first()

    if not attendance:
        return jsonify({
            "checked_in": False
        })

    return jsonify({
    "checked_in": True,

    "check_in":
        attendance.check_in.isoformat(),

    "lunch_break":
        attendance.lunch_break,

    "tea_break":
        attendance.tea_break,

    "lunch_start":
        attendance.lunch_start.isoformat()
        if attendance.lunch_start
        else None,

    "tea_start":
        attendance.tea_start.isoformat()
        if attendance.tea_start
        else None,

    "lunch_minutes":
        attendance.lunch_minutes or 0,

    "tea_minutes":
        attendance.tea_minutes or 0,

    "total_break_minutes":
        attendance.total_break_minutes or 0
})




@attendance_bp.route(
    "/lunch-break",
    methods=["POST"]
)
def lunch_break():

    try:

        data = request.json

        attendance = Attendance.query.filter_by(
            user_id=data["user_id"],
            check_out=None
        ).order_by(
            Attendance.id.desc()
        ).first()

        if not attendance:
            return jsonify({
                "success": False,
                "error": "Attendance not found"
            }), 404

        action = data.get("action")

        if action == "start":

            attendance.lunch_break = True

            attendance.lunch_start = datetime.now()

        elif action == "stop":

            attendance.lunch_break = False

            attendance.lunch_end = datetime.now()

        if (
             attendance.lunch_start and
             attendance.lunch_end
            ):
                attendance.lunch_minutes = int(
        (
            attendance.lunch_end -
            attendance.lunch_start
        ).total_seconds() / 60
    )

        attendance.total_break_minutes = (
            (attendance.lunch_minutes or 0) +
            (attendance.tea_minutes or 0)
        )

        db.session.commit()

        return jsonify({
            "success": True
        })

    except Exception as e:

        print("LUNCH BREAK ERROR:", str(e))

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@attendance_bp.route(
    "/tea-break",
    methods=["POST"]
)
def tea_break():

    try:

        data = request.json

        attendance = Attendance.query.filter_by(
            user_id=data["user_id"],
            check_out=None
        ).order_by(
            Attendance.id.desc()
        ).first()

        if not attendance:
            return jsonify({
                "success": False,
                "error": "Attendance not found"
            }), 404

        action = data.get("action")

        if action == "start":

            attendance.tea_break = True

            attendance.tea_start = datetime.now()

        elif action == "stop":
            attendance.tea_break = False

            attendance.tea_end = datetime.now()
        if (
            attendance.tea_start and
            attendance.tea_end
            ):
              attendance.tea_minutes = int(
        (
            attendance.tea_end -
            attendance.tea_start
        ).total_seconds() / 60
    )

        attendance.total_break_minutes = (
            (attendance.lunch_minutes or 0) +
            (attendance.tea_minutes or 0)
        )

        db.session.commit()

        return jsonify({
            "success": True
        })

    except Exception as e:

        print("TEA BREAK ERROR:", str(e))

        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@attendance_bp.route("/history/<int:user_id>")
def attendance_history(user_id):

    records = Attendance.query.filter_by(
        user_id=user_id
    ).order_by(
        Attendance.id.desc()
    ).all()

    result = []

    for record in records:
        result.append({
    "id": record.id,
    "date":
    record.attendance_date.strftime("%Y-%m-%d")
    if record.attendance_date
    else "-",
    "checkIn": record.check_in.strftime("%I:%M %p") if record.check_in else "-",
    "checkOut": record.check_out.strftime("%I:%M %p") if record.check_out else "-",
    "workingHours": record.total_hours,
    "lunchMinutes": record.lunch_minutes,
    "teaMinutes": record.tea_minutes,
    "totalBreak": record.total_break_minutes,
    "status": record.status
})

    return jsonify(result)


@attendance_bp.route("/", methods=["GET"])
def get_attendance():

    today = datetime.now().date()

    employees = Employee.query.all()

    attendance_list = []

    for employee in employees:

        attendance = Attendance.query.filter_by(
            user_id=employee.user_id,
            attendance_date=today
        ).first()

        if attendance:

            status = attendance.status or "Present"

            check_in = (
                attendance.check_in.strftime("%H:%M:%S")
                if attendance.check_in
                else "-"
            )

            check_out = (
                attendance.check_out.strftime("%H:%M:%S")
                if attendance.check_out
                else "-"
            )

            total_hours = attendance.total_hours

        else:

            status = "Absent"
            check_in = "-"
            check_out = "-"
            total_hours = 0

        attendance_list.append({
            "user_id": employee.user_id,
            "employee_name": f"{employee.first_name} {employee.last_name}",
            "department": employee.department,
            "designation": employee.designation,
            "check_in": check_in,
            "check_out": check_out,
            "total_hours": total_hours,
            "attendance_date": str(today),
            "status": status,

            "shift_timing": (
                attendance.shift_timing
                if attendance and attendance.shift_timing
                else employee.shift_timing or "General Shift"
            )
        })

    return jsonify(attendance_list)

@attendance_bp.route("/generate-daily-attendance")
def generate_daily_attendance():

    today = date.today()

    users = User.query.filter_by(
        is_active=True
    ).all()

    count = 0

    for user in users:

        existing = Attendance.query.filter_by(
            user_id=user.id,
            attendance_date=today
        ).first()

        if not existing:

            attendance = Attendance(
                user_id=user.id,
                attendance_date=today,
                status="Absent"
            )

            db.session.add(attendance)
            count += 1

    db.session.commit()

    return jsonify({
        "success": True,
        "records_created": count
    })



@attendance_bp.route(
    "/weekly",
    methods=["GET"]
)
def get_weekly_attendance():

    try:

        result = []

        employees = Employee.query.all()

        # Today first
        for i in range(7):

            current_date = (
                date.today() - timedelta(days=i)
            )

            for employee in employees:

                attendance = Attendance.query.filter_by(
                    user_id=employee.user_id,
                    attendance_date=current_date
                ).first()

                result.append({

                    "employee_name":
                        f"{employee.first_name} {employee.last_name}",

                    "team":
                        employee.department
                        if employee.department
                        else "-",

                    "date":
                        current_date.strftime("%d-%m-%Y"),

                    "check_in":
                        attendance.check_in.strftime("%I:%M %p")
                        if attendance and attendance.check_in
                        else "-",

                    "check_out":
                        attendance.check_out.strftime("%I:%M %p")
                        if attendance and attendance.check_out
                        else "-",

                    "total_hours":
                        attendance.total_hours
                        if attendance
                        else "-",

                    "status":
                        attendance.status
                        if attendance
                        else "Absent",

                    "shift_timing":
                        attendance.shift_timing
                        if attendance and attendance.shift_timing
                        else (
                            employee.shift_timing
                            or "General Shift"
                        )
                })

        return jsonify(result)

    except Exception as e:

        return jsonify({
            "error": str(e)
        }), 500


@attendance_bp.route(
    "/monthly",
    methods=["GET"]
)
def get_monthly_attendance():

    try:

        result = []

        employees = Employee.query.all()

        # Last 30 days - newest first
        for i in range(30):

            current_date = (
                date.today() - timedelta(days=i)
            )

            for employee in employees:

                attendance = Attendance.query.filter_by(
                    user_id=employee.user_id,
                    attendance_date=current_date
                ).first()

                result.append({

                    "employee_name":
                        f"{employee.first_name} {employee.last_name}",

                    "team":
                        employee.department
                        if employee.department
                        else "-",

                    "date":
                        current_date.strftime("%d-%m-%Y"),

                    "check_in":
                        attendance.check_in.strftime("%I:%M %p")
                        if attendance and attendance.check_in
                        else "-",

                    "check_out":
                        attendance.check_out.strftime("%I:%M %p")
                        if attendance and attendance.check_out
                        else "-",

                    "total_hours":
                        attendance.total_hours
                        if attendance
                        else "-",

                    "status":
                        attendance.status
                        if attendance
                        else "Absent",

                    "shift_timing":
                        attendance.shift_timing
                        if attendance and attendance.shift_timing
                        else (
                            employee.shift_timing
                            or "General Shift"
                        )
                })

        return jsonify(result)

    except Exception as e:

        return jsonify({
            "error": str(e)
        }), 500
    
@attendance_bp.route(
    "/export-monthly",
    methods=["GET"]
)
def export_monthly_attendance():

    try:

        wb = Workbook()

        ws = wb.active

        ws.title = "Attendance Report"

        # =====================================
        # STYLES
        # =====================================

        purple_fill = PatternFill(
            fill_type="solid",
            fgColor="B58CE5"
        )

        yellow_fill = PatternFill(
            fill_type="solid",
            fgColor="F7F1A0"
        )

        white_font = Font(
            bold=True,
            color="FFFFFF",
            size=12
        )

        bold_font = Font(
            bold=True,
            size=12
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # =====================================
        # TITLE
        # =====================================

        ws.merge_cells("A1:H1")

        ws["A1"] = "ATTENDANCE REPORT"

        ws["A1"].fill = purple_fill

        ws["A1"].font = Font(
            bold=True,
            size=16,
            color="FFFFFF"
        )

        ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # =====================================
        # MONTH HEADER
        # =====================================

        ws.merge_cells("A2:H2")

        ws["A2"] = (
            f"Attendance Summary "
            f"{date.today().strftime('%B %Y')}"
        )

        ws["A2"].fill = purple_fill

        ws["A2"].font = white_font

        ws["A2"].alignment = Alignment(
            horizontal="center"
        )

        # =====================================
        # DATE RANGE
        # =====================================

        start_date = date.today() - timedelta(days=30)

        ws.merge_cells("A3:H3")

        ws["A3"] = (
            f"Date Range : "
            f"{start_date.strftime('%d-%b-%Y')} "
            f"to "
            f"{date.today().strftime('%d-%b-%Y')}"
        )

        ws["A3"].fill = yellow_fill

        ws["A3"].font = bold_font

        ws["A3"].alignment = Alignment(
            horizontal="center"
        )

        # =====================================
        # COLUMN HEADERS
        # =====================================

        headers = [

            "S.No",
            "Emp Code",
            "Emp Name",
            "D.O.J",
            "Department",
            "Days Payable",
            "Days Worked",
            "Shift"

        ]

        for col_num, header in enumerate(
            headers,
            start=1
        ):

            cell = ws.cell(
                row=5,
                column=col_num
            )

            cell.value = header

            cell.fill = purple_fill

            cell.font = white_font

            cell.border = thin_border

            cell.alignment = Alignment(
                horizontal="center"
            )

        # =====================================
        # EMPLOYEE DATA
        # =====================================

        employees = Employee.query.all()

        row = 6

        for index, employee in enumerate(
            employees,
            start=1
        ):

            attendance_records = Attendance.query.filter_by(
                user_id=employee.user_id
            ).all()

            days_worked = len([
                a
                for a in attendance_records
                if a.status == "Present"
            ])

            ws.cell(
                row=row,
                column=1
            ).value = index

            ws.cell(
                row=row,
                column=2
            ).value = (
                employee.employee_id
                if hasattr(employee, "employee_id")
                else employee.user_id
            )

            ws.cell(
                row=row,
                column=3
            ).value = (
                f"{employee.first_name} "
                f"{employee.last_name}"
            )

            ws.cell(
                row=row,
                column=4
            ).value = (
                str(employee.joining_date)
                if hasattr(employee, "joining_date")
                and employee.joining_date
                else ""
            )

            ws.cell(
                row=row,
                column=5
            ).value = (
                employee.department
                if employee.department
                else "-"
            )

            ws.cell(
                row=row,
                column=6
            ).value = 30

            ws.cell(
                row=row,
                column=7
            ).value = days_worked

            ws.cell(
                row=row,
                column=8
            ).value = (
                employee.shift_timing
                if employee.shift_timing
                else "General Shift"
            )

            for col in range(1, 9):

                ws.cell(
                    row=row,
                    column=col
                ).border = thin_border

            row += 1

        # =====================================
        # AUTO WIDTH
        # =====================================

        for column_cells in ws.columns:

            length = max(
                len(str(cell.value))
                if cell.value
                else 0
                for cell in column_cells
            )

            ws.column_dimensions[
                get_column_letter(
                    column_cells[0].column
                )
            ].width = length + 5

        # =====================================
        # FILTER
        # =====================================

        ws.auto_filter.ref = (
            f"A5:H{row}"
        )

        # =====================================
        # SAVE FILE
        # =====================================

        output = BytesIO()

        wb.save(output)

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="Attendance_Report.xlsx",
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        )

    except Exception as e:

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500